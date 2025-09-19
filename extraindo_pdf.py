
# -*- coding: utf-8 -*-
"""
Gera um Excel organizado (OS, Itens, Totais) a partir de um PDF de OS/Contrato.

- Compatível com Python 3.7+ (inclui 3.13.x)
- Requer: pymupdf, pandas, openpyxl
- Formata a aba Itens como Tabela do Excel (com filtros) e aplica formatos numéricos.

Uso básico:
    python gera_excel_os.py --input frota_4100408.pdf

Uso com saída personalizada:
    python gera_excel_os.py --input frota_4100408.pdf --output resultado.xlsx
"""

import re
import argparse
from decimal import Decimal, InvalidOperation
from pathlib import Path

import fitz  # PyMuPDF
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# -------------------- Utilidades --------------------

def br_to_decimal(s: str) -> Decimal:
    """Converte strings pt-BR '1.234,56' para Decimal('1234.56')."""
    if not s:
        return Decimal('0')
    s = s.strip().replace('.', '').replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')


def autosize_columns(ws, max_width=60):
    """Ajusta largura de colunas com base no conteúdo."""
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


# -------------------- Extração do PDF --------------------

def extract_from_pdf(pdf_path: Path) -> dict:
    """Extrai dados estruturados do PDF fornecido."""
    with fitz.open(str(pdf_path)) as doc:
        text = "\n".join(page.get_text("text") for page in doc)

    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    full = "\n".join(lines)

    # Nº OS
    numero_os = None
    m = re.search(r"\bN[ºo]\s*(\d{5,})", full)
    if m:
        numero_os = m.group(1)
    else:
        # fallback: número logo acima de "Cliente:"
        for i, l in enumerate(lines):
            if l == "Cliente:" or l.startswith("Cliente:"):
                for j in range(i - 1, max(-1, i - 5), -1):
                    if re.fullmatch(r"\d{5,}", lines[j]):
                        numero_os = lines[j]
                        break
                break

    # Helper: pega próximo valor após um rótulo
    def next_value_after(label: str):
        for i, l in enumerate(lines):
            if l.startswith(label):
                for j in range(i + 1, min(i + 6, len(lines))):
                    if not lines[j].endswith(":"):
                        return lines[j]
        return None

    emissao = next_value_after("Emissão:")

    # Cliente
    cliente_codigo = None
    try:
        idx = lines.index("Cliente:")
        if idx + 1 < len(lines) and re.fullmatch(r"[A-Z0-9]+", lines[idx + 1]):
            cliente_codigo = lines[idx + 1]
    except ValueError:
        pass

    cliente_nome, cliente_email = None, None
    for l in lines:
        if "@" in l and not l.lower().startswith("e-mail"):
            em = re.search(r"[\w\.-]+@[\w\.-]+", l)
            if em:
                cliente_email = em.group(0)
                cliente_nome = l.replace(cliente_email, "").strip(" -")
                cliente_nome = re.sub(" +", " ", cliente_nome)
                break

    endereco = None
    for l in lines:
        if " - " in l and re.search(r"/[A-Z]{2}\b", l):
            endereco = l
            break

    telefones = None
    for l in lines:
        if re.search(r"\(\d{2}\)", l):
            telefones = l
            break

    cnpj = None
    m = re.search(r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b", full)
    if m:
        cnpj = m.group(0)

    rg_ie = None
    for l in lines:
        if re.fullmatch(r"\d{10,15}", l):
            rg_ie = l
            break

    # Veículo
    frota = next_value_after("Frota:")
    placa = next_value_after("Placa:")

    # KM (evita confundir com telefones)
    km = None
    for i, l in enumerate(lines):
        if l.startswith("KM:"):
            for j in range(i + 1, min(i + 6, len(lines))):
                cand = lines[j]
                if cand.endswith(":"):
                    continue
                if re.search(r"\d", cand) and not re.search(r"\(\d{2}\)", cand):
                    km = cand
                    break
            break

    # Observações
    observ_geral = None
    for i, l in enumerate(lines):
        if l.lower().startswith("observações geral"):
            if i + 1 < len(lines):
                observ_geral = lines[i + 1]
            break

    # Itens (máquina de estados de 7 passos)
    items = []
    try:
        start = lines.index("Referência") + 1
    except ValueError:
        start = 0

    state = "qtd"
    cur, i = {}, start
    while i < len(lines):
        l = lines[i]
        if l.startswith("Total Bruto:"):
            break

        if state == "qtd":
            if re.fullmatch(r"\d{1,3},\d{2}", l):
                cur = {"qtd": l}
                state = "unit"
            i += 1
            continue

        if state == "unit":
            if re.fullmatch(r"\d{1,3}(?:\.\d{3})*,\d{2}", l):
                cur["unitario"] = l
                state = "total"
            i += 1
            continue
        if state == "total":
            if re.fullmatch(r"\d{1,3}(?:\.\d{3})*,\d{2}", l):
                cur["total"] = l
                state = "descricao"
            i += 1
            continue

        if state == "descricao":
            cur["descricao"] = l
            state = "ncm"
            i += 1
            continue

        if state == "ncm":
            if re.fullmatch(r"\d{4}\.\d{2}\.\d{2}", l):
                cur["ncm"] = l
                state = "referencia"
            else:
                # NCM ausente → descarta bloco atual e recomeça
                state = "qtd"
            i += 1
            continue

        if state == "referencia":
            if re.fullmatch(r"[\d\.]+", l):
                cur["referencia"] = l
                state = "produto"
            i += 1
            continue

        if state == "produto":
            if re.fullmatch(r"\d{3,}", l):
                cur["produto"] = l
                items.append(cur)
                cur = {}
                state = "qtd"
            i += 1
            continue

    # Totais
    bruto = sum(br_to_decimal(it.get("total")) for it in items)

    desconto = Decimal("0.00")
    for i, l in enumerate(lines):
        if l.startswith("Total Desconto:"):
            for j in range(i + 1, min(i + 8, len(lines))):
                if re.fullmatch(r"\d{1,3}(?:\.\d{3})*,\d{2}", lines[j]):
                    desconto = br_to_decimal(lines[j])
                    break
            break

    liquido = bruto - desconto

    result = {
        "arquivo": str(pdf_path.name),
        "numero_os": numero_os,
        "emissao": emissao,
        "cliente": {
            "codigo": cliente_codigo,
            "nome": cliente_nome,
            "email": cliente_email,
            "cpf_cnpj": cnpj,
            "rg_ie": rg_ie,
            "endereco": endereco,
            "telefones": telefones,
        },
        "veiculo": {"frota": frota, "placa": placa, "km": km},
        "observacoes": {"geral": observ_geral},
        "itens": items,
        "totais": {"bruto": bruto, "desconto": desconto, "liquido": liquido},
    }
    return result


# -------------------- Excel --------------------

def export_to_excel(res: dict, xlsx_path: Path) -> None:
    """Gera um Excel com abas OS, Itens (tabela), Totais e formatações."""
    # Aba OS
    os_row = {
        "Nº OS": res.get("numero_os"),
        "Emissão": res.get("emissao"),
        "Cliente Código": res["cliente"].get("codigo"),
        "Cliente Nome": res["cliente"].get("nome"),
        "Cliente E-mail": res["cliente"].get("email"),
        "CNPJ": res["cliente"].get("cpf_cnpj"),
        "RG/IE": res["cliente"].get("rg_ie"),
        "Endereço": res["cliente"].get("endereco"),
        "Telefones": res["cliente"].get("telefones"),
        "Frota": res["veiculo"].get("frota"),
        "Placa": res["veiculo"].get("placa"),
        "KM": res["veiculo"].get("km"),
        "Observações": res["observacoes"].get("geral"),
    }
    os_df = pd.DataFrame([os_row])

    # Aba Itens
    items_df = pd.DataFrame(res["itens"])
    have_items = not items_df.empty
    if have_items:
        cols_order = ["produto", "descricao", "ncm", "referencia", "qtd", "unitario", "total"]
        items_df = items_df.reindex(columns=cols_order)
        # Converte para numéricos reais
        items_df["Qtd"] = items_df["qtd"].apply(br_to_decimal).astype(float)
        items_df["Unitário"] = items_df["unitario"].apply(br_to_decimal).astype(float)
        items_df["Total"] = items_df["total"].apply(br_to_decimal).astype(float)
        # Renomeia para saída
        items_out = items_df[["produto", "descricao", "ncm", "referencia", "Qtd", "Unitário", "Total"]].copy()
        items_out.columns = ["Produto", "Descrição", "NCM", "Referência", "Qtd", "Unitário", "Total"]
    else:
        items_out = pd.DataFrame(columns=["Produto", "Descrição", "NCM", "Referência", "Qtd", "Unitário", "Total"])

    # Totais
    totais_out = pd.DataFrame([
        {"Tipo": "Bruto", "Valor": float(res["totais"]["bruto"])},
        {"Tipo": "Desconto", "Valor": float(res["totais"]["desconto"])},
        {"Tipo": "Líquido", "Valor": float(res["totais"]["liquido"])},
    ])

    # Grava e formata
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        os_df.to_excel(writer, sheet_name="OS", index=False)
        items_out.to_excel(writer, sheet_name="Itens", index=False)
        totais_out.to_excel(writer, sheet_name="Totais", index=False)

        wb = writer.book

        # --- OS ---
        ws_os = wb["OS"]
        autosize_columns(ws_os, max_width=60)

        # --- Itens ---
        ws_it = wb["Itens"]
        ws_it.freeze_panes = "A2"  # congela cabeçalho
        # Formatação numérica (Qtd, Unitário, Total = colunas E, F, G)
        for row in range(2, ws_it.max_row + 1):
            ws_it[f"E{row}"].number_format = "0.00"
            ws_it[f"F{row}"].number_format = "R$ #,##0.00"
            ws_it[f"G{row}"].number_format = "R$ #,##0.00"
        autosize_columns(ws_it, max_width=50)
        # Cria Tabela com estilo
        if ws_it.max_row >= 2 and ws_it.max_column >= 1:
            last_col_letter = get_column_letter(ws_it.max_column)
            ref = f"A1:{last_col_letter}{ws_it.max_row}"
            table = Table(displayName="TabelaItens", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws_it.add_table(table)

        # --- Totais ---
        ws_tt = wb["Totais"]
        for row in range(2, ws_tt.max_row + 1):
            ws_tt[f"B{row}"].number_format = "R$ #,##0.00"
        autosize_columns(ws_tt, max_width=30)


# -------------------- Main (CLI) --------------------

def main():
    ap = argparse.ArgumentParser(description="Gera Excel organizado (OS, Itens, Totais) a partir de PDF.")
    ap.add_argument("--input", "-i", required=False, default="frota_4100408.pdf",
                    help="Arquivo PDF de entrada (padrão: frota_4100408.pdf)")
    ap.add_argument("--output", "-o", required=False,
                    help="Arquivo XLSX de saída (padrão: mesmo nome do PDF, com .xlsx)")
    args = ap.parse_args()

    pdf_path = Path(args.input)
    if not pdf_path.exists():
        raise FileNotFoundError(f"Arquivo PDF não encontrado: {pdf_path}")

    xlsx_path = Path(args.output) if args.output else pdf_path.with_suffix(".xlsx")

    res = extract_from_pdf(pdf_path)
    export_to_excel(res, xlsx_path)

    print(f"✅ Excel gerado com sucesso: {xlsx_path.resolve()}")


if __name__ == "__main__":
    main()
